# src/extract/serial_reader.py
from openpyxl import load_workbook
import os
import pandas as pd
import numpy as np
from src.transform.state_manager import get_row_hash

def parse_serial_sheet(file_path: str, serial_hashes: dict = None):
    """
    Parses the 'serial-full' sheet of the Excel workbook using openpyxl
    to ensure exact data types and formats are preserved (especially leading zeros).
    
    Args:
        file_path: Path to the target excel file.
        serial_hashes: Optional dictionary mapping section names to sets of already processed hashes.
        
    Returns:
        sections_to_write: list of dicts: [{"name": section_name, "rows": [row_dict, ...]}]
        all_serial_hashes: dict mapping section_name to list of all hashes currently in the sheet.
    """
    wb = load_workbook(file_path, data_only=True)
    if "serial-full" not in wb.sheetnames:
        return [], {}
        
    sheet = wb["serial-full"]
    sections_to_write = []
    all_serial_hashes = {}
    
    # Helper to clean values preserving string formatting
    def clean_serial_val(val):
        if val is None or (isinstance(val, float) and np.isnan(val)):
            return None
        if isinstance(val, bool):
            return val
        if isinstance(val, (int, float)):
            if isinstance(val, float) and val.is_integer():
                return int(val)
            return val
        if isinstance(val, str):
            val = val.strip()
            if val == "(null)":
                return None
            val = val.replace('\xa0', '').strip()
            if not val:
                return None
            if len(val) >= 2 and val.startswith('"') and val.endswith('"'):
                val = val[1:-1]
            if val.lower() == 'true':
                return True
            if val.lower() == 'false':
                return False
            # Keep string with leading zeroes as string
            if val.isdigit() and val.startswith('0') and len(val) > 1:
                return val
            try:
                return int(val)
            except ValueError:
                pass
            try:
                f = float(val)
                if f.is_integer():
                    return int(f)
                return f
            except ValueError:
                pass
            return val
        return val

    r = 1
    max_row = sheet.max_row
    current_section = None
    
    while r <= max_row:
        cell_val = sheet.cell(row=r, column=1).value
        # Check if cell starts a new section
        if cell_val is not None and isinstance(cell_val, str) and cell_val.strip().startswith("Record"):
            if current_section:
                sections_to_write.append(current_section)
                
            section_name = cell_val.strip()
            current_section = {
                "name": section_name,
                "headers": [],
                "rows": []
            }
            all_serial_hashes[section_name] = []
            
            # Read headers from the next row (r+1)
            headers_row = r + 1
            current_headers = []
            c = 1
            while True:
                h_val = sheet.cell(row=headers_row, column=c).value
                if h_val is None:
                    break
                current_headers.append(str(h_val).strip())
                c += 1
            current_section["headers"] = current_headers
            
            r += 2
            continue
            
        if current_section:
            # Check if row is empty
            row_empty = True
            for col_idx in range(1, len(current_section["headers"]) + 1):
                if sheet.cell(row=r, column=col_idx).value is not None:
                    row_empty = False
                    break
                    
            if row_empty:
                sections_to_write.append(current_section)
                current_section = None
                r += 1
                continue
                
            # Read data row
            row_dict = {}
            for col_idx, header in enumerate(current_section["headers"]):
                val = sheet.cell(row=r, column=col_idx + 1).value
                cleaned = clean_serial_val(val)
                if cleaned is not None:
                    row_dict[header] = cleaned
            
            if row_dict:
                row_hash = get_row_hash(row_dict)
                section_name = current_section["name"]
                
                # Record the hash as existing in current sheet
                all_serial_hashes[section_name].append(row_hash)
                
                # Check if it is a new row
                is_new = True
                if serial_hashes is not None:
                    processed_set = serial_hashes.get(section_name, set())
                    if row_hash in processed_set:
                        is_new = False
                
                if is_new:
                    current_section["rows"].append(row_dict)
                    
        r += 1
        
    if current_section:
        sections_to_write.append(current_section)
        
    # Filter out sections that have no new rows (if running incrementally)
    if serial_hashes is not None:
        sections_to_write = [s for s in sections_to_write if len(s["rows"]) > 0]
        
    return sections_to_write, all_serial_hashes
